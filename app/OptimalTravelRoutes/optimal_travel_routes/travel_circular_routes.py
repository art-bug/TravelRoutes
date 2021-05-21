from openpyxl import load_workbook
from collections import defaultdict

def get_worksheets(excel_workbook_name):
    excel_workbook = load_workbook(excel_workbook_name)
    return list(map(excel_workbook.__getitem__, excel_workbook.sheetnames))

def get_distance_matrix(distance_matrix_sheet, empty_value=0):
    header = next(distance_matrix_sheet.iter_rows(min_col=2, values_only=True))
    splitted_city_labels = [(*city_label.rsplit(', ', 1)[::-1], index) for index, city_label in enumerate(header)]
    country_cities = defaultdict(list)

    for country, city, city_index in splitted_city_labels:
        country_cities[country].append((city, city_index))

    iter_rows = distance_matrix_sheet.iter_rows(min_row=2, min_col=2, values_only=True)
    prepared_iter_rows = map(lambda row:
                                map(lambda distance:
                                        distance if distance not in ('', None) else empty_value,
                                    row),
                            iter_rows)

    distance_matrix = list(map(list, prepared_iter_rows))

    return country_cities, distance_matrix

def get_city_populations(populations_sheet):
    iter_rows = populations_sheet.iter_rows(min_row=2, values_only=True)

    city_populations = {city.rsplit(', ', 1)[0]:population for city, population in iter_rows}

    return city_populations

def cond_distances(city_row, max_distance):
    return ((city_index, distance) for city_index, distance in enumerate(city_row) if distance > 0 and distance <= max_distance)

def any_circular_routes(distance_matrix, city_index, max_distance):
    return len(list(cond_distances(distance_matrix[city_index], max_distance))) >= 2 
 
def travel_circular_routes(distance_matrix, city_index, max_distance, travel_days):
    circular_routes = []

    if not any_circular_routes(distance_matrix, city_index, max_distance):
        return circular_routes

    def descendants(current_city):
        return cond_distances(distance_matrix[current_city], max_distance)

    def dfs_cycles(distance_matrix, start=city_index, current=city_index, required_length=travel_days, path=[city_index], visited=defaultdict(bool)):
        nonlocal circular_routes

        visited[current] = True

        if len(path) == required_length:
            visited[current] = False

            if start in map(lambda descendant: descendant[0], descendants(current)):
                circular_routes.append(path + [start])

            return

        for next_city, distance in descendants(current):
            if not visited.get(next_city, False):
                dfs_cycles(distance_matrix, start, next_city, required_length, path + [next_city], visited)
                visited[next_city] = True
 
        visited[current] = False

        return

    dfs_cycles(distance_matrix)

    return circular_routes

def set_score(route, city_populations: dict):
    return (route, sum(map(city_populations.__getitem__, route[1:-1])))

def sort_circular_routes(circular_routes, ascending=False):
    return sorted(circular_routes, key=lambda route: route[1], reverse=not ascending)

